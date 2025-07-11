import json
import os
import time
from urllib.parse import urljoin

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager


class BaseCrawler:
    """
    爬虫基类，提供通用的爬取和数据处理功能
    """

    def __init__(self, brand_name, data_dir="data"):
        """
        初始化爬虫基类

        Args:
            brand_name: 品牌名称，用于文件命名
            data_dir: 数据目录
        """
        # 基本属性
        self.brand_name = brand_name
        self.data_dir = data_dir
        self.base_url = None
        self.start_urls = []

        # 记录开始信息
        print(f"开始爬取 {brand_name} 数据")

        # 确保数据目录存在
        if not os.path.exists(self.data_dir):
            os.makedirs(self.data_dir)

        # 配置Chrome
        chrome_options = Options()
        chrome_options.add_argument("--headless")
        chrome_options.add_argument("--disable-gpu")
        chrome_options.add_argument("--window-size=1920,1080")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument(
            "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36")

        # 初始化WebDriver
        self.driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)
        self.wait = WebDriverWait(self.driver, 10)

        # 存储爬取的数据
        self.all_products = []

    def close(self):
        """关闭WebDriver"""
        if self.driver:
            self.driver.quit()

    def get_links_from_page(self, url, selector):
        """
        获取页面中匹配选择器的所有链接

        Args:
            url: 要爬取的URL
            selector: CSS选择器

        Returns:
            匹配的链接列表
        """
        links = []
        try:
            self.driver.get(url)
            time.sleep(2)  # 等待页面加载

            elements = self.driver.find_elements(By.CSS_SELECTOR, selector)
            for element in elements:
                href = element.get_attribute('href')
                if href:
                    # 处理相对URL
                    if not href.startswith(('http://', 'https://')):
                        href = urljoin(self.base_url, href)
                    links.append(href)

            return links
        except Exception as e:
            print(f"获取链接时出错 {url}: {str(e)}")
            return []

    def extract_product_details(self, url):
        """
        从产品页面提取产品详情，子类必须实现此方法

        Args:
            url: 产品页面URL

        Returns:
            产品数据字典或None（如果提取失败）
        """
        raise NotImplementedError("子类必须实现extract_product_details方法")

    def process_category_page(self, url):
        """
        处理类别页面并处理分页，子类必须实现此方法

        Args:
            url: 类别页面URL

        Returns:
            该类别下所有产品链接的列表
        """
        raise NotImplementedError("子类必须实现process_category_page方法")

    def run(self):
        """
        运行爬虫，处理所有起始URL并提取产品详情
        """
        skipped_urls = []

        try:
            all_product_links = []

            # 处理每个起始页面
            for start_url in self.start_urls:
                print(f"开始处理类别页面: {start_url}")
                product_links = self.process_category_page(start_url)
                all_product_links.extend(product_links)

            print(f"所有类别页面共找到 {len(all_product_links)} 个产品链接")

            # 处理每个产品链接
            for i, product_link in enumerate(all_product_links):
                print(f"处理产品 {i + 1}/{len(all_product_links)}: {product_link}")
                product_data = self.extract_product_details(product_link)
                if product_data:
                    self.all_products.append(product_data)
                else:
                    skipped_urls.append(product_link)

            # 保存结果
            if self.all_products:
                self.process_and_save_data()

                # 保存被跳过的URL
                if skipped_urls:
                    with open(os.path.join(self.data_dir, f'{self.brand_name}_skipped_urls.txt'), 'w',
                              encoding='utf-8') as f:
                        for url in skipped_urls:
                            f.write(f"{url}\n")

                print(f"爬取完成。产品总数: {len(self.all_products)}, 跳过URL数: {len(skipped_urls)}")
            else:
                print("未采集到任何有效数据")

        except Exception as e:
            print(f"运行爬虫时出错: {str(e)}")

        finally:
            self.close()

    def prepare_data(self):
        """
        准备基本数据和规格参数数据

        Returns:
            (products_df, specs_df): 包含产品信息和规格参数的DataFrame
        """
        try:
            # 准备基本产品信息数据
            basic_data = []
            for product in self.all_products:
                basic_data.append({
                    'id': product['id'],
                    'name': product['name'],
                    'url': product['url']
                })

            # 创建产品基本信息DataFrame
            products_df = pd.DataFrame(basic_data)

            # 准备规格参数数据（展平规格参数）
            specs_data = []
            for product in self.all_products:
                for param_name, param_value in product['specifications'].items():
                    specs_data.append({
                        'product_id': product['id'],
                        'product_name': product['name'],
                        'param_name': param_name,
                        'param_value': param_value
                    })

            # 创建规格参数DataFrame
            specs_df = pd.DataFrame(specs_data)

            return products_df, specs_df

        except Exception as e:
            print(f"准备数据时出错: {str(e)}")
            return None, None

    def create_excel_file(self, specs_df, product_name_map, output_file):
        """
        创建格式化的Excel文件，包含四列：产品ID、产品名称、参数名称、参数值

        Args:
            specs_df: 规格参数DataFrame
            product_name_map: 产品ID到名称的映射字典
            output_file: 输出文件路径
        """
        # 获取所有唯一的产品ID
        product_ids = specs_df['product_id'].unique()
        print(f"Excel格式化: 发现 {len(product_ids)} 个唯一产品ID")

        # 创建一个新的工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "产品规格参数"

        # 添加标题行
        ws.append(["产品ID", "产品名称", "参数名称", "参数值"])

        # 设置单元格边框样式
        thin_border = Side(border_style="thin", color="000000")
        border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)

        # 遍历每个产品ID
        row_index = 2  # 从第2行开始（第1行是标题）
        for product_id in product_ids:
            # 获取当前产品ID的所有行
            product_rows = specs_df[specs_df['product_id'] == product_id]

            # 获取产品名称
            product_name = product_name_map.get(product_id, "未知产品")

            # 记录当前产品的开始行
            start_row = row_index

            # 添加产品的所有参数行
            for _, row in product_rows.iterrows():
                ws.append([product_id, product_name, row['param_name'], row['param_value']])

                # 设置单元格对齐方式和边框
                for col in range(1, 5):
                    cell = ws.cell(row=row_index, column=col)
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    cell.border = border

                row_index += 1

            # 合并产品ID列和产品名称列的单元格（如果有多行）
            if start_row != row_index - 1:
                ws.merge_cells(f'A{start_row}:A{row_index - 1}')  # 合并产品ID列
                ws.merge_cells(f'B{start_row}:B{row_index - 1}')  # 合并产品名称列

                # 设置合并后单元格的对齐方式
                merged_cell_id = ws.cell(row=start_row, column=1)
                merged_cell_id.alignment = Alignment(horizontal='left', vertical='center')

                merged_cell_name = ws.cell(row=start_row, column=2)
                merged_cell_name.alignment = Alignment(horizontal='left', vertical='center')

        # 调整列宽
        ws.column_dimensions['A'].width = 20  # 产品ID
        ws.column_dimensions['B'].width = 30  # 产品名称
        ws.column_dimensions['C'].width = 30  # 参数名称
        ws.column_dimensions['D'].width = 50  # 参数值

        # 保存文件
        wb.save(output_file)
        print(f"Excel格式化: 结果已保存到 {output_file}")

    def create_pivot_csv(self, specs_df, output_file):
        """
        创建透视表格式的CSV文件，每行一个产品，所有参数作为列

        Args:
            specs_df: 规格参数DataFrame
            output_file: 输出文件路径
        """
        # 透视表转换 - 将参数名称作为列
        pivot_df = specs_df.pivot_table(
            index=['product_id', 'product_name'],
            columns='param_name',
            values='param_value',
            aggfunc='first'
        ).reset_index()

        # 保存处理后的透视表CSV文件
        pivot_df.to_csv(output_file, index=False, encoding='utf-8-sig')
        print(f"透视表格式: 结果已保存到 {output_file}")

    def create_json_file(self, specs_df, output_file):
        """
        创建JSON格式文件，每个产品为一个对象，包含参数列表

        Args:
            specs_df: 规格参数DataFrame
            output_file: 输出文件路径
        """
        # 组装为JSON结构
        result = []
        for pid, group in specs_df.groupby('product_id'):
            name = group['product_name'].iloc[0] if not group['product_name'].isnull().all() else ""
            params = []
            for _, row in group.iterrows():
                params.append({
                    "paramName": row['param_name'],
                    "param": row['param_value']
                })
            result.append({
                "product_id": pid,
                "product_name": name,
                "params": params
            })

        # 保存为JSON
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(result, f, ensure_ascii=False, indent=2)

        print(f"JSON格式: 结果已保存到 {output_file}")

    def process_and_save_data(self):
        """
        处理爬取的数据并保存为指定的三种格式
        """
        # 1. 准备数据
        products_df, specs_df = self.prepare_data()
        if products_df is None or specs_df is None:
            print("数据准备失败，无法保存")
            return

        # 2. 创建产品ID到名称的映射
        product_name_map = dict(zip(products_df['id'], products_df['name']))

        # 3. 生成Excel格式文件（带合并单元格）
        excel_file = os.path.join(self.data_dir, f'{self.brand_name}_specifications_formatted.xlsx')
        self.create_excel_file(specs_df, product_name_map, excel_file)

        # 4. 生成透视表CSV
        pivot_file = os.path.join(self.data_dir, f'{self.brand_name}_specifications_pivot.csv')
        self.create_pivot_csv(specs_df, pivot_file)

        # 5. 生成JSON文件
        json_file = os.path.join(self.data_dir, f'{self.brand_name}_specifications.json')
        self.create_json_file(specs_df, json_file)

        # 6. 汇总所有生成的文件
        output_files = {
            'excel_file': excel_file,
            'pivot_file': pivot_file,
            'json_file': json_file
        }

        print("\n数据处理完成，已生成以下文件:")
        for key, file_path in output_files.items():
            print(f"- {key}: {file_path}")
